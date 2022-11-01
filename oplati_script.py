import pandas as pd
import vk_api
from openpyxl import load_workbook
from my_token import my_token
'''
cols = [0,1,5]
flash_gr_pup = pd.read_excel('./Оплаты.xlsx', usecols = cols)
'''
#vk_api
group_id = 215643440
vk_session = vk_api.VkApi(token=my_token)
session_api = vk_session.get_api()
res_api = vk_session.method('groups.getLongPollServer',{'group_id':group_id})
def send_message(user_id, message):
  message_load = {'user_id': user_id, 'random_id': 0, **message}
  vk_session.method('messages.send', message_load)
def find_amount(group_id):
  group_getMembers = {'group_id': group_id, 'offset': 0}
  return vk_session.method('groups.getMembers', group_getMembers)
print("Введи свою группу в формате {буква}{номер}") #Example a8
group_num = str(input())
group_num = group_num.lower()
pup_amount = find_amount(group_id)
k = pup_amount['count']
end = 2000
vk_ids = pup_amount['items']
#send_message(560830328, {'message': "привет"})
#Excel info
flash_wb = load_workbook('./Оплаты.xlsx')
op_sheet = flash_wb.get_sheet_by_name('Лист1')
name_price = {}
for i in range(1, end):
  if op_sheet.cell(row = i, column = 1).value == group_num:
      if len(name_price)==0:
        end = i + k
      if op_sheet.cell(row = i, column = 11).value == 0 and op_sheet.cell(row = i, column = 10).value != 1:
        if int(op_sheet.cell(row = i, column = 15).value[op_sheet.cell(row = i, column = 15).value.find("id") + 2:]) in vk_ids:
          vk_id = int(op_sheet.cell(row = i, column = 15).value[op_sheet.cell(row = i, column = 15).value.find("id") + 2:])
          if op_sheet.cell(row = i, column = 12).value != 'слив' and op_sheet.cell(row = i, column = 12).value != 'кик':
            if 'уш' not in str(op_sheet.cell(row = i, column = 12).value):
              if op_sheet.cell(row = i, column = 10).value is None:
                    op_needed = int(op_sheet.cell(row = i, column = 9).value)
              else:
                    op_needed = int(op_sheet.cell(row = i, column = 9).value) - int(op_sheet.cell(row = i, column = 10).value)
              name_price[vk_id] = op_needed
for op in name_price:
  send_message(op, {'message': "привет, наконец наступило время оплат. В этом месяце тебе надо заплатить " +
                    str(name_price[op]) + ". Сделай это, пожалуйста побыстрее, чтобы я потом лишний раз тебя не тревожил"})
print(res_api)
